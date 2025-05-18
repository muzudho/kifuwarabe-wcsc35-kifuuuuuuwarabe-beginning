import cshogi
import openpyxl as xl

from ...models.layer_o1o0o1o0_japanese import JapaneseMoveModel


class XlSearchModule:
    """Excel を用いた探索部。
    """


    _PATH_TO_XL_SEARCH_WORKSHEET = './temp/search_part_work.xlsx'


    def __init__(self, gymnasium):
        """［初期化］

        Parameters
        ----------
        gymnasium : GymnasiumModel
            ［体育館］
            記憶は全部この中に詰め込めます。
        """
        self._gymnasium = gymnasium


    def start_search(self, move_list):

        # ワークブックを新規生成
        wb = xl.Workbook()

        # ワークシート
        ws = wb['Sheet']

        ws['A1'].value = '1階合法手'
        for index, move in enumerate(move_list):

            # 指し手のUSI表記に、独自形式を併記。
            ws[f"A{index + 2}"] = JapaneseMoveModel.from_move(
                    move    = move,
                    cap_pt  = self._gymnasium.table.piece_type(sq=cshogi.move_to(move)),
                    is_mars = False,
                    is_gote = self._gymnasium.table.is_gote).stringify()

        ws[f"A{len(move_list) + 2}"].value = '(ここまで)'

        wb.save(filename=XlSearchModule._PATH_TO_XL_SEARCH_WORKSHEET)
