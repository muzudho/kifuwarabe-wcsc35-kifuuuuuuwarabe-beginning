import cshogi
import openpyxl as xl

from openpyxl.styles import PatternFill, Font
from ...models.layer_o1o0 import constants
from ...models.layer_o1o0o1o0_japanese import JapaneseMoveModel


class XlSearchModel:
    """Excel を用いた探索部。
    """


    _PATH_TO_XL_SEARCH_WORKSHEET = './temp/search_part_work.xlsx'


    def __init__(self, gymnasium):
        """［初期化］

        Parameters
        ----------
        gymnasium : GymnasiumModel
            ［体育館］
            記憶は全部この中に詰め込めます。
        """
        self._gymnasium = gymnasium

        # 色
        self._HEADER_BLACK = '404040'
        self._HEADER_WHITE = 'F2F2F2'

        # フォント
        self._HEADER_FONT = Font(size=12.0, color=self._HEADER_WHITE)

        self._wb = None
        self._ws = None


    _TERMINATION = {
        None: None,
        'GAME_OVER' : constants.out_of_termination_state_const.GAME_OVER,
        'NYUGYOKU_WIN' : constants.out_of_termination_state_const.NYUGYOKU_WIN,
        #'MATE_IN_1_MOVE' : constants.out_of_termination_state_const.MATE_IN_1_MOVE,
    }

    def get_termination_state(self, number_of_moves, row_th):
        """終端の状況を取得。

        Parameters
        ----------
        number_of_moves : int
            何手目か。
        row_th : int
            行番号。
        """

        column_name = f"{number_of_moves}階終端"
        column_letter = self._get_column_letter_of_column_name(column_name=column_name)   # Excel シートを横に探索。［n階終端］列を探す。
        termination_str = self._ws[f"{column_letter}{row_th}"].value
        return XlSearchModel._TERMINATION[termination_str]


    def render_1st_move(self, move_list):
        """１手目を描画。
        """
        try:
            # ワークブックを新規生成
            self._wb = xl.Workbook()

            # ワークシート
            self._ws = self._wb['Sheet']

            cell = self._ws['A1']
            cell.value = '削除'
            cell.fill = PatternFill(patternType='solid', fgColor=self._HEADER_BLACK)
            cell.font = self._HEADER_FONT

            cell = self._ws['B1']
            cell.value = '1階合法手'
            cell.fill = PatternFill(patternType='solid', fgColor=self._HEADER_BLACK)
            cell.font = self._HEADER_FONT

            cell = self._ws['C1']
            cell.value = '1階終端'
            cell.fill = PatternFill(patternType='solid', fgColor=self._HEADER_BLACK)
            cell.font = self._HEADER_FONT

            row_th = 2

            # 指す前
            self._ws[f"B{row_th}"].value = '(指す前)'

            if self._gymnasium.table.is_game_over():
                """投了局面時。
                """
                self._ws[f"C{row_th}"].value = 'GAME_OVER'
                return

            if self._gymnasium.table.is_nyugyoku():
                """入玉宣言勝ち局面時。
                """
                self._ws[f"C{row_th}"].value = 'NYUGYOKU_WIN'
                return

            # 一手詰めを詰める
            if not self._gymnasium.table.is_check():
                """自玉に王手がかかっていない時で"""

                if (matemove := self._gymnasium.table.mate_move_in_1ply()):
                    """一手詰めの指し手があれば、それを取得"""
                    self._ws[f"C{row_th}"].value = 'MATE_IN_1_MOVE'
                    return

            row_th += 1

            # 指し手
            for index, move in enumerate(move_list):

                # 指し手のUSI表記に、独自形式を併記。
                self._ws[f"B{row_th}"].value = JapaneseMoveModel.from_move(
                        move    = move,
                        cap_pt  = self._gymnasium.table.piece_type(sq=cshogi.move_to(move)),
                        is_mars = False,
                        is_gote = self._gymnasium.table.is_gote).stringify()

                row_th += 1

        except Exception as ex:
            print(ex)


    def save_worksheet(self):
        self._wb.save(filename=XlSearchModel._PATH_TO_XL_SEARCH_WORKSHEET)



    ####################
    # MARK: サブルーチン
    ####################

    def _get_column_letter_of_column_name(self, column_name):
        """列名を指定したら、列番号を返したい。

        Returns
        -------
        該当が無ければナン。
        """

        row_th = 1
        column_th = 1

        while True:
            column_letter = xl.utils.get_column_letter(column_th)
            cell = self._ws[f"{column_letter}{row_th}"]
            if cell.value == column_name:
                return column_letter
            elif cell.value == '':
                return None
            
            column_th += 1
