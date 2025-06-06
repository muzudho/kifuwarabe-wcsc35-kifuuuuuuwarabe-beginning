import cshogi
import os
import openpyxl as xl
import pyxlart as xa
import re

from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment


from ....models.layer_o1o_8o0_str import StringResourcesModel
from ....models.layer_o1o0 import PieceModel, SquareModel, TurnModel
from .xs_utils import XsUtils


class XsBoardView():


    def __init__(self):
        # 色
        BLACK = '000000'
        BACKGROUND_COLOR = 'F2DCDB'
        # 盤の色
        BOARD_MARS_LIGHT_COLOR = 'B7DEE8'
        BOARD_MARS_SOFT_COLOR = '92CDDC'
        BOARD_RIVER_LIGHT_COLOR = 'DDD9C4'
        BOARD_RIVER_SOFT_COLOR = 'C4BD97'
        BOARD_EARTH_LIGHT_COLOR = 'D8E4BC'
        BOARD_EARTH_SOFT_COLOR = 'C4D79B'
        HEADER_1_COLOR = 'FCD5B4'
        HEADER_2_COLOR = 'FDE9D9'
        TITLE_COLOR = '4F81BD'
        # ハイライトと影の色
        MARS_HIGHLIGHT_1 = 'DAEEF3'
        MARS_HIGHLIGHT_2 = 'B7DEE8'
        MARS_SHADOW_1 = '31869B'
        MARS_SHADOW_2 = '215967'
        RIVER_HIGHLIGHT_1 = 'EEECE1'
        RIVER_HIGHLIGHT_2 = 'DDD9C4'
        RIVER_SHADOW_1 = '948A54'
        RIVER_SHADOW_2 = '494529'   # 暗すぎる。 1D1B10
        EARTH_HIGHLIGHT_1 = 'EBF1DE'
        EARTH_HIGHLIGHT_2 = 'D8E4BC'
        EARTH_SHADOW_1 = '76933C'
        EARTH_SHADOW_2 = '4F6228'
        FILE_NUMBER_COLOR = 'B1A0C7'
        LEFT_RANK_NUMBER_COLOR = '948A54'
        RIGHT_RANK_NUMBER_COLOR = '95B3D7'

        # フォント
        self._NEXT_LABEL_FONT = Font(size=12.0, color='31869B')
        self._NEXT_VALUE_FONT = Font(size=12.0, bold=True, color='60497A')
        self._TITLE_FONT = Font(size=16.0, color=TITLE_COLOR)
        self._SMALL_TITLE_FONT = Font(size=6.0, color=TITLE_COLOR)
        self._FILE_NUMBER_FONT = Font(size=20.0, color=FILE_NUMBER_COLOR)
        self._LEFT_RANK_NUMBER_FONT = Font(size=20.0, color=LEFT_RANK_NUMBER_COLOR)
        self._RIGHT_RANK_NUMBER_FONT = Font(size=20.0, color=RIGHT_RANK_NUMBER_COLOR)
        self._MARS_HANDS_NUMBER_FONT = Font(size=20.0, color=MARS_SHADOW_2)
        self._EARTH_HANDS_NUMBER_FONT = Font(size=20.0, color=EARTH_SHADOW_2)

        # 罫線の要素
        self._thin_black_side = Side(style='thin', color=BLACK)
        self._thick_black_side = Side(style='thick', color=BLACK)
        board_frame_mars_highlight1_side = Side(style='thick', color=MARS_HIGHLIGHT_1)
        board_frame_mars_highlight2_side = Side(style='thick', color=MARS_HIGHLIGHT_2)
        board_frame_river_highlight1_side = Side(style='thick', color=RIVER_HIGHLIGHT_1)
        #board_frame_river_highlight2_side = Side(style='thick', color=RIVER_HIGHLIGHT_2)
        board_frame_earth_highlight1_side = Side(style='thick', color=EARTH_HIGHLIGHT_1)
        board_frame_earth_highlight2_side = Side(style='thick', color=EARTH_HIGHLIGHT_2)
        #board_frame_earth_soft_side = Side(style='thick', color=BOARD_EARTH_SOFT_COLOR)
        board_frame_mars_shadow1_side = Side(style='thick', color=MARS_SHADOW_1)
        board_frame_mars_shadow2_side = Side(style='thick', color=MARS_SHADOW_2)
        board_frame_river_shadow1_side = Side(style='thick', color=RIVER_SHADOW_1)
        #board_frame_river_shadow2_side = Side(style='thick', color=RIVER_SHADOW_2)
        board_frame_earth_shadow1_side = Side(style='thick', color=EARTH_SHADOW_1)
        board_frame_earth_shadow2_side = Side(style='thick', color=EARTH_SHADOW_2)

        # 罫線
        self._board_cell_border = Border(left=self._thin_black_side, right=self._thin_black_side, top=self._thin_black_side, bottom=self._thin_black_side)
        self._board_top_left_border = Border(left=self._thick_black_side, top=self._thick_black_side)
        self._board_top_border = Border(top=self._thick_black_side)
        self._board_top_right_border = Border(right=self._thick_black_side, top=self._thick_black_side)
        self._board_left_border = Border(left=self._thick_black_side)
        self._board_right_border = Border(right=self._thick_black_side)
        self._board_bottom_left_border = Border(left=self._thick_black_side, bottom=self._thick_black_side)
        self._board_bottom_border = Border(bottom=self._thick_black_side)
        self._board_bottom_right_border = Border(right=self._thick_black_side, bottom=self._thick_black_side)
        # 盤の枠の罫線
        self._border_of_frame_of_mars_light_top_left = Border(left=board_frame_mars_highlight1_side, top=board_frame_mars_highlight1_side)
        self._border_of_frame_of_mars_light_top = Border(top=board_frame_mars_highlight1_side)
        self._border_of_frame_of_mars_light_top_right = Border(top=board_frame_mars_highlight1_side, right=board_frame_mars_shadow1_side)
        self._border_of_frame_of_mars_light_left = Border(left=board_frame_mars_highlight1_side)
        self._border_of_frame_of_mars_light_right = Border(right=board_frame_mars_shadow1_side)
        self._border_of_frame_of_mars_soft_left = Border(left=board_frame_mars_highlight2_side)
        self._border_of_frame_of_mars_soft_right = Border(right=board_frame_mars_shadow2_side)
        self._border_of_frame_of_mars_light_bottom_left = Border(left=board_frame_mars_highlight1_side, bottom=board_frame_mars_shadow1_side)
        self._border_of_frame_of_mars_light_bottom = Border(bottom=board_frame_mars_shadow1_side)
        self._border_of_frame_of_mars_light_bottom_right = Border(right=board_frame_mars_shadow1_side, bottom=board_frame_mars_shadow1_side)
        self._border_of_frame_of_river_light_left = Border(left=board_frame_river_highlight1_side)
        self._border_of_frame_of_river_light_right = Border(right=board_frame_river_shadow1_side)
        self._border_of_frame_of_earth_light_top_left = Border(left=board_frame_earth_highlight1_side, top=board_frame_earth_highlight1_side)
        self._border_of_frame_of_earth_light_top = Border(top=board_frame_earth_highlight1_side)
        self._border_of_frame_of_earth_light_top_right = Border(top=board_frame_earth_highlight1_side, right=board_frame_earth_shadow1_side)
        self._border_of_frame_of_earth_light_left = Border(left=board_frame_earth_highlight1_side)
        self._border_of_frame_of_earth_light_right = Border(right=board_frame_earth_shadow1_side)
        self._border_of_frame_of_earth_soft_left = Border(left=board_frame_earth_highlight2_side)
        self._border_of_frame_of_earth_soft_right = Border(right=board_frame_earth_shadow2_side)
        self._border_of_frame_of_earth_light_bottom_left = Border(left=board_frame_earth_highlight1_side, bottom=board_frame_earth_shadow1_side)
        self._border_of_frame_of_earth_light_bottom = Border(bottom=board_frame_earth_shadow1_side)
        self._border_of_frame_of_earth_light_bottom_right = Border(right=board_frame_earth_shadow1_side, bottom=board_frame_earth_shadow1_side)

        # フィル
        self._background_fill = PatternFill(patternType='solid', fgColor=BACKGROUND_COLOR)
        self._board_mars_light_fill = PatternFill(patternType='solid', fgColor=BOARD_MARS_LIGHT_COLOR)
        self._board_mars_soft_fill = PatternFill(patternType='solid', fgColor=BOARD_MARS_SOFT_COLOR)
        self._board_river_light_fill = PatternFill(patternType='solid', fgColor=BOARD_RIVER_LIGHT_COLOR)
        self._board_river_soft_fill = PatternFill(patternType='solid', fgColor=BOARD_RIVER_SOFT_COLOR)
        self._board_earth_light_fill = PatternFill(patternType='solid', fgColor=BOARD_EARTH_LIGHT_COLOR)
        self._board_earth_soft_fill = PatternFill(patternType='solid', fgColor=BOARD_EARTH_SOFT_COLOR)
        self._header_1_fill = PatternFill(patternType='solid', fgColor=HEADER_2_COLOR)
        self._header_2_fill = PatternFill(patternType='solid', fgColor=HEADER_1_COLOR)

        # 寄せ
        #   horizontal は 'distributed', 'fill', 'general', 'center', 'centerContinuous', 'justify', 'right', 'left' のいずれかから選ぶ
        #   vertical は 'center', 'top', 'bottom', 'justify', 'distributed' のいずれかから選ぶ
        self._left_top_alignment = Alignment(horizontal='left', vertical='top')
        self._left_center_alignment = Alignment(horizontal='left', vertical='center')
        self._center_center_alignment = Alignment(horizontal='center', vertical='center')
        self._right_center_alignment = Alignment(horizontal='right', vertical='center')


    def render(self, gymnasium):
        """描画。
        """

        # ワークブックを新規生成
        wb = xl.Workbook()

        # ワークシート
        ws = wb['Sheet']

        # 方眼紙を作成します。
        xa.GraphPaperRenderer.render(
                left_th = 1,
                top_th  = 1,
                width   = 100,
                height  = 100,
                ws      = ws)

        self._render_background(ws)                     # 背景を描画。
        self._render_next_label_1(ws)                   # 何手目ラベル１。
        self._render_next_value(ws, gymnasium)          # 何手目値。
        self._render_moves_label_2(ws)                  # 何手目ラベル２。
        self._render_color(ws, gymnasium)               # 手番の描画。
        self._render_repetition_label(ws)               # 局面反復回数ラベル。
        self._render_repetition_value(ws)               # 局面反復回数値。
        self._render_mars_hands(ws, gymnasium)          # 一段目側の持ち駒の描画。
        self._render_earth_hands(ws, gymnasium)         # 九段目側の持ち駒の描画。
        self._render_board_background_except_squares(ws, gymnasium)    # 盤の背景を描画。マスを除く。
        self._render_board_each_square(ws, gymnasium)     # 盤の各マスを描画。
        self._render_title(ws)                          # タイトルを描画。

        # ワークブック保存
        gymnasium.exshell.save_workbook(wb=wb)

        # エクセル開く
        gymnasium.exshell.open_virtual_display()

        line = input('何か入力してください。例： y\n')
        print() # 空行

        # エクセル閉じる
        gymnasium.exshell.close_virtual_display()


    def _render_background(self, ws):
        """背景を描画。
        """
        start_row_th = 1
        end_row_th = 28
        for y_th in range(start_row_th, end_row_th + 1):
            for column_letter in xa.ColumnLetterRange(start='A', end='AK'):
                row_th = y_th
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._background_fill


    def _render_next_label_1(self, ws):
        """何手目ラベル１。
        """
        cell = ws['C2']
        cell.value = 'Next'
        cell.font = self._NEXT_LABEL_FONT
        cell.fill = self._header_2_fill
        cell.alignment = self._right_center_alignment
        ws.merge_cells('C2:D2')


    def _render_next_value(self, ws, gymnasium):
        """何手目値。
        """
        cell = ws['E2']
        cell.value = gymnasium.table.move_number
        cell.font = self._NEXT_VALUE_FONT
        cell.fill = self._header_1_fill
        cell.alignment = self._center_center_alignment
        ws.merge_cells('E2:F2')


    def _render_moves_label_2(self, ws):
        """何手目ラベル２。
        """
        cell = ws['G2']
        cell.value = 'move(s)'
        cell.font = self._NEXT_LABEL_FONT
        cell.fill = self._header_2_fill
        cell.alignment = self._left_center_alignment
        ws.merge_cells('G2:J2')


    def _render_color(self, ws, gymnasium):
        """手番の描画。
        """
        cell = ws['K2']
        cell.value = TurnModel.code(gymnasium.table.turn).capitalize()
        cell.font = self._NEXT_VALUE_FONT
        cell.fill = self._header_1_fill
        cell.alignment = self._center_center_alignment
        ws.merge_cells('K2:L2')


    def _render_repetition_label(self, ws):
        """局面反復回数ラベル。
        """
        cell = ws['M2']
        cell.value = 'Repetition'
        cell.font = self._NEXT_LABEL_FONT
        cell.fill = self._header_2_fill
        cell.alignment = self._right_center_alignment
        ws.merge_cells('M2:P2')


    def _render_repetition_value(self, ws):
        """局面反復回数値。
        """
        cell = ws['Q2']
        cell.value = '_'
        cell.font = self._NEXT_VALUE_FONT
        cell.fill = self._header_1_fill
        cell.alignment = self._left_center_alignment


    def _render_mars_hands(self, ws, gymnasium):
        """一段目側の持ち駒の描画。
        """
        # 駒台を塗り潰し
        row_th = 5
        for column_letter in xa.ColumnLetterRange(start='C', end='G'):
            cell = ws[f"{column_letter}{row_th}"]
            cell.fill = self._board_mars_light_fill
        for row_th in range(6, 8):
            for column_letter in xa.ColumnLetterRange(start='C', end='G'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_mars_soft_fill
        for row_th in range(8, 10):
            for column_letter in xa.ColumnLetterRange(start='C', end='G'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_mars_light_fill
        for row_th in range(10, 12):
            for column_letter in xa.ColumnLetterRange(start='C', end='G'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_mars_soft_fill
        for row_th in range(12, 14):
            for column_letter in xa.ColumnLetterRange(start='C', end='G'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_mars_light_fill
        for row_th in range(14, 16):
            for column_letter in xa.ColumnLetterRange(start='C', end='G'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_mars_soft_fill
        for row_th in range(16, 18):
            for column_letter in xa.ColumnLetterRange(start='C', end='G'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_mars_light_fill
        for row_th in range(18, 20):
            for column_letter in xa.ColumnLetterRange(start='C', end='G'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_mars_soft_fill
        row_th = 20
        for column_letter in xa.ColumnLetterRange(start='C', end='G'):
            cell = ws[f"{column_letter}{row_th}"]
            cell.fill = self._board_mars_light_fill

        # 枠の罫線
        # 上辺 5行目
        row_th = 5
        ws[f"C{row_th}"].border = self._border_of_frame_of_mars_light_top_left
        for column_letter in xa.ColumnLetterRange(start='D', end='F'):
            ws[f"{column_letter}{row_th}"].border = self._border_of_frame_of_mars_light_top
        ws[f"F{row_th}"].border = self._border_of_frame_of_mars_light_top_right
        # Soft部 6～7行目
        for row_th in range(6, 8):
            ws[f"C{row_th}"].border = self._border_of_frame_of_mars_soft_left
            ws[f"F{row_th}"].border = self._border_of_frame_of_mars_soft_right
        # Light部 8～9行目
        for row_th in range(8, 10):
            ws[f"C{row_th}"].border = self._border_of_frame_of_mars_light_left
            ws[f"F{row_th}"].border = self._border_of_frame_of_mars_light_right
        # Soft部 10～11行目
        for row_th in range(10, 12):
            ws[f"C{row_th}"].border = self._border_of_frame_of_mars_soft_left
            ws[f"F{row_th}"].border = self._border_of_frame_of_mars_soft_right
        # Light部 12～13行目
        for row_th in range(12, 14):
            ws[f"C{row_th}"].border = self._border_of_frame_of_mars_light_left
            ws[f"F{row_th}"].border = self._border_of_frame_of_mars_light_right
        # Soft部 14～16行目
        for row_th in range(14, 16):
            ws[f"C{row_th}"].border = self._border_of_frame_of_mars_soft_left
            ws[f"F{row_th}"].border = self._border_of_frame_of_mars_soft_right
        # Light部 16～18行目
        for row_th in range(16, 18):
            ws[f"C{row_th}"].border = self._border_of_frame_of_mars_light_left
            ws[f"F{row_th}"].border = self._border_of_frame_of_mars_light_right
        # Soft部 18～20行目
        for row_th in range(18, 20):
            ws[f"C{row_th}"].border = self._border_of_frame_of_mars_soft_left
            ws[f"F{row_th}"].border = self._border_of_frame_of_mars_soft_right
        # 下辺 20行目
        row_th = 20
        ws[f"C{row_th}"].border = self._border_of_frame_of_mars_light_bottom_left
        for column_letter in xa.ColumnLetterRange(start='D', end='F'):
            ws[f"{column_letter}{row_th}"].border = self._border_of_frame_of_mars_light_bottom
        ws[f"F{row_th}"].border = self._border_of_frame_of_mars_light_bottom_right

        for row_th in range(6, 19, 2):  # セル結合
            ws.merge_cells(f"E{row_th}:F{row_th+1}")
            cell = ws[f"E{row_th}"]
            cell.font = self._MARS_HANDS_NUMBER_FONT
            cell.alignment = self._center_center_alignment

        # 後手の持ち駒の数のリスト
        w_hand = gymnasium.table.pieces_in_hand[1]
        #
        # NOTE 元の画像サイズで貼り付けられるわけではないの、何でだろう？ 60x60pixels の画像にしておくと、90x90pixels のセルに合う？
        #
        # TODO 📖 [PythonでExcelファイルに画像を挿入する/列の幅を調整する](https://qiita.com/kaba_san/items/b231a41891ebc240efc7)
        # 難しい
        #
        input_data = [
            ('C6', cshogi.WHITE, cshogi.PAWN, 'E6', w_hand[6]),
            ('C8', cshogi.WHITE, cshogi.LANCE, 'E8', w_hand[5]),
            ('C10', cshogi.WHITE, cshogi.KNIGHT, 'E10', w_hand[4]),
            ('C12', cshogi.WHITE, cshogi.SILVER, 'E12', w_hand[3]),
            ('C14', cshogi.WHITE, cshogi.GOLD, 'E14', w_hand[2]),
            ('C16', cshogi.WHITE, cshogi.BISHOP, 'E16', w_hand[1]),
            ('C18', cshogi.WHITE, cshogi.ROOK, 'E18', w_hand[0]),
        ]
        for (cell_number, color, pt, value_cell_address, number_of_piece) in input_data:
            if 0 < number_of_piece:
                XsUtils.render_piece_1(ws=ws, cell_address=cell_number, color=color, pt=pt)
                ws[value_cell_address].value = number_of_piece


    def _render_earth_hands(self, ws, gymnasium):
        """九段目側の持ち駒の描画。
        """
        # 駒台を塗り潰し
        row_th = 9
        for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
            cell = ws[f"{column_letter}{row_th}"]
            cell.fill = self._board_earth_light_fill
        for row_th in range(10, 12):
            for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_earth_soft_fill
        for row_th in range(12, 14):
            for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_earth_light_fill
        for row_th in range(14, 16):
            for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_earth_soft_fill
        for row_th in range(16, 18):
            for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_earth_light_fill
        for row_th in range(18, 20):
            for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_earth_soft_fill
        for row_th in range(20, 22):
            for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_earth_light_fill
        for row_th in range(22, 24):
            for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
                cell = ws[f"{column_letter}{row_th}"]
                cell.fill = self._board_earth_soft_fill
        row_th = 24
        for column_letter in xa.ColumnLetterRange(start='AE', end='AI'):
            cell = ws[f"{column_letter}{row_th}"]
            cell.fill = self._board_earth_light_fill

        # 枠の罫線
        # 上辺 9行目
        row_th = 9
        ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_light_top_left
        for column_letter in xa.ColumnLetterRange(start='AF', end='AH'):
            ws[f"{column_letter}{row_th}"].border = self._border_of_frame_of_earth_light_top
        ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_light_top_right
        # Soft部 10～11行目
        for row_th in range(10, 12):
            ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_soft_left
            ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_soft_right
        # Light部 12～13行目
        for row_th in range(12, 14):
            ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_light_left
            ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_light_right
        # Soft部 14～15行目
        for row_th in range(14, 16):
            ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_soft_left
            ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_soft_right
        # Light部 16～17行目
        for row_th in range(16, 18):
            ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_light_left
            ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_light_right
        # Soft部 18～19行目
        for row_th in range(18, 20):
            ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_soft_left
            ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_soft_right
        # Light部 20～21行目
        for row_th in range(20, 22):
            ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_light_left
            ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_light_right
        # Soft部 22～23行目
        for row_th in range(22, 24):
            ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_soft_left
            ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_soft_right
        # 下辺 24行目
        row_th = 24
        ws[f"AE{row_th}"].border = self._border_of_frame_of_earth_light_bottom_left
        for column_letter in xa.ColumnLetterRange(start='AF', end='AH'):
            ws[f"{column_letter}{row_th}"].border = self._border_of_frame_of_earth_light_bottom
        ws[f"AH{row_th}"].border = self._border_of_frame_of_earth_light_bottom_right

        for row_th in range(10, 23, 2):  # セル結合
            ws.merge_cells(f"AG{row_th}:AH{row_th+1}")
            cell = ws[f"AG{row_th}"]
            cell.font = self._EARTH_HANDS_NUMBER_FONT
            cell.alignment = self._center_center_alignment

        # 先手、後手の持ち駒の数のリスト
        b_hand = gymnasium.table.pieces_in_hand[0]
        #
        # NOTE 元の画像サイズで貼り付けられるわけではないの、何でだろう？ 60x60pixels の画像にしておくと、90x90pixels のセルに合う？
        #
        # TODO 📖 [PythonでExcelファイルに画像を挿入する/列の幅を調整する](https://qiita.com/kaba_san/items/b231a41891ebc240efc7)
        # 難しい
        #
        input_data = [
            ('AE10', cshogi.BLACK, cshogi.PAWN, 'AG10', b_hand[6]),
            ('AE12', cshogi.BLACK, cshogi.LANCE, 'AG12', b_hand[5]),
            ('AE14', cshogi.BLACK, cshogi.KNIGHT, 'AG14', b_hand[4]),
            ('AE16', cshogi.BLACK, cshogi.SILVER, 'AG16', b_hand[3]),
            ('AE18', cshogi.BLACK, cshogi.GOLD, 'AG18', b_hand[2]),
            ('AE20', cshogi.BLACK, cshogi.BISHOP, 'AG20', b_hand[1]),
            ('AE22', cshogi.BLACK, cshogi.ROOK, 'AG22', b_hand[0]),
        ]
        for (cell_number, color, pt, value_cell_address, number_of_piece) in input_data:
            if 0 < number_of_piece:
                XsUtils.render_piece_1(ws=ws, cell_address=cell_number, color=color, pt=pt)
                ws[value_cell_address].value = number_of_piece


    def _render_board_background_except_squares(self, ws, gymnasium):
        """盤の背景を描画。マスを除く。
        """
        # 枠の辺を塗り潰し
        # 上辺
        for row_th in range(4, 6):
            for column_letter in xa.ColumnLetterRange(start='H', end='AD'):
                ws[f"{column_letter}{row_th}"].fill = self._board_mars_light_fill

        # 下辺
        for row_th in range(24, 26):
            for column_letter in xa.ColumnLetterRange(start='H', end='AD'):
                ws[f"{column_letter}{row_th}"].fill = self._board_earth_light_fill

        # 左辺
        for row_th in range(6, 12):
            for column_letter in xa.ColumnLetterRange(start='H', end='J'):
                ws[f"{column_letter}{row_th}"].fill = self._board_mars_light_fill
        for row_th in range(12, 18):
            for column_letter in xa.ColumnLetterRange(start='H', end='J'):
                ws[f"{column_letter}{row_th}"].fill = self._board_river_light_fill
        for row_th in range(18, 24):
            for column_letter in xa.ColumnLetterRange(start='H', end='J'):
                ws[f"{column_letter}{row_th}"].fill = self._board_earth_light_fill

        # 右辺
        for row_th in range(6, 12):
            for column_letter in xa.ColumnLetterRange(start='AB', end='AD'):
                ws[f"{column_letter}{row_th}"].fill = self._board_mars_light_fill
        for row_th in range(12, 18):
            for column_letter in xa.ColumnLetterRange(start='AB', end='AD'):
                ws[f"{column_letter}{row_th}"].fill = self._board_river_light_fill
        for row_th in range(18, 24):
            for column_letter in xa.ColumnLetterRange(start='AB', end='AD'):
                ws[f"{column_letter}{row_th}"].fill = self._board_earth_light_fill

        # 筋の番号
        for index, column_letter in enumerate(xa.ColumnLetterRange(start='J', end='AA', step=2)):
            next_column_letter = xa.ColumnLetterLogic.add(column_letter, 1)
            ws.merge_cells(f"{column_letter}4:{next_column_letter}5")
            cell = ws[f"{column_letter}4"]
            cell.value = (9 - index) * 10
            cell.font = self._FILE_NUMBER_FONT
            cell.alignment = self._center_center_alignment

        # 左側の段の番号
        for index, row_th in enumerate(range(6, 23, 2)):
            ws.merge_cells(f"H{row_th}:I{row_th+1}")
            cell = ws[f"H{row_th}"]
            cell.value = f"{StringResourcesModel.small_alphabet_list()[index+1]}"
            cell.font = self._LEFT_RANK_NUMBER_FONT
            cell.alignment = self._center_center_alignment

        # 右側の段の番号
        for index, row_th in enumerate(range(6, 23, 2)):
            ws.merge_cells(f"AB{row_th}:AC{row_th+1}")
            cell = ws[f"AB{row_th}"]
            cell.value = index + 1
            cell.font = self._RIGHT_RANK_NUMBER_FONT
            cell.alignment = self._center_center_alignment

        # 盤の枠の罫線
        # 上辺 4行目
        row_th = 4
        ws[f"H{row_th}"].border = self._border_of_frame_of_mars_light_top_left
        for column_letter in xa.ColumnLetterRange(start='I', end='AC'):
            ws[f"{column_letter}{row_th}"].border = self._border_of_frame_of_mars_light_top
        ws[f"AC{row_th}"].border = self._border_of_frame_of_mars_light_top_right
        # Mars部 5～11行目
        for row_th in range(5, 12):
            ws[f"H{row_th}"].border = self._border_of_frame_of_mars_light_left
            ws[f"AC{row_th}"].border = self._border_of_frame_of_mars_light_right
        # River部 12～17行目
        for row_th in range(12, 18):
            ws[f"H{row_th}"].border = self._border_of_frame_of_river_light_left
            ws[f"AC{row_th}"].border = self._border_of_frame_of_river_light_right
        # Earth部 18～24行目
        for row_th in range(18, 25):
            ws[f"H{row_th}"].border = self._border_of_frame_of_earth_light_left
            ws[f"AC{row_th}"].border = self._border_of_frame_of_earth_light_right
        # 下辺 25行目
        row_th = 25
        ws[f"H{row_th}"].border = self._border_of_frame_of_earth_light_bottom_left
        for column_letter in xa.ColumnLetterRange(start='I', end='AC'):
            ws[f"{column_letter}{row_th}"].border = self._border_of_frame_of_earth_light_bottom
        ws[f"AC{row_th}"].border = self._border_of_frame_of_earth_light_bottom_right


    def _render_board_each_square(self, ws, gymnasium):
        """盤の各マスを描画。
        """

        # マスの色指定
        square_fills = []

        for i in range(0,27):
            if i % 2 == 0:
                square_fills.append(self._board_mars_soft_fill)
            else:
                square_fills.append(self._board_mars_light_fill)

        for i in range(0,27):
            if i % 2 == 0:
                square_fills.append(self._board_river_light_fill)
            else:
                square_fills.append(self._board_river_soft_fill)

        for i in range(0,27):
            if i % 2 == 0:
                square_fills.append(self._board_earth_soft_fill)
            else:
                square_fills.append(self._board_earth_light_fill)

        # 盤の各マス
        index = 0   # 通し番号
        for row_th in range(6, 23, 2):
            for column_letter in xa.ColumnLetterRange(start='J', end='AA', step=2):
                # セル設定
                cell = ws[f"{column_letter}{row_th}"]
                cell.border = self._board_cell_border

                cell.fill = square_fills[index]

                next_column_letter = xa.ColumnLetterLogic.add(column_letter, 1)

                # セル結合
                ws.merge_cells(f"{column_letter}{row_th}:{next_column_letter}{row_th+1}")
                index += 1

        # 盤の枠を太線にします。セル結合を考えず描きます。
        ws['J6'].border = self._board_top_left_border
        ws['AA6'].border = self._board_top_right_border
        ws['J23'].border = self._board_bottom_left_border
        ws['AA23'].border = self._board_bottom_right_border
        for column_letter in xa.ColumnLetterRange(start='K', end='AA'):
            cell = ws[xa.CellAddressModel.from_code(f"{column_letter}6").to_code()]
            cell.border = xa.BorderLogic.add(
                    base        = cell.border,
                    addition    = self._board_top_border)

            cell = ws[xa.CellAddressModel.from_code(f"{column_letter}23").to_code()]
            cell.border = xa.BorderLogic.add(
                    base        = cell.border,
                    addition    = self._board_bottom_border)

        for row_th in range(7, 23):
            cell = ws[f"J{row_th}"]
            cell.border = xa.BorderLogic.add(
                    base        = cell.border,
                    addition    = self._board_left_border)

            cell = ws[f"AA{row_th}"]
            cell.border = xa.BorderLogic.add(
                    base        = cell.border,
                    addition    = self._board_right_border)

        # 駒を描画
        masu = 91
        for column_letter in xa.ColumnLetterRange(start='J', end='AB', step=2):
            for row_th in range(6, 24, 2):
                sq = SquareModel.from_masu(masu=masu).sq
                piece = gymnasium.table.piece(sq)
                color = PieceModel.turn(piece)
                pt = cshogi.piece_to_piece_type(piece)
                #print(f"{masu=} {sq=}")
                XsUtils.render_piece_2(
                        ws=ws,
                        sq=sq,
                        color=color,
                        pt=pt)
                masu -= 10
            # 11 を 82 に変換
            # 12 を 72 に変換
            next_dan = masu % 10 + 1
            next_suji = 9
            masu = next_suji * 10 + next_dan
            #print(f"{next_suji=} {next_dan=} {masu=}")


    def _render_title(self, ws):
        """タイトルを描画。
        """

        input_data = [
            ('L27', 'B'),
            ('M27', 'i'),
            ('N27', 'g'),

            ('P27', 'f'),
            ('Q27', 'o'),
            ('R27', 'r'),
            ('S27', 'e'),
            ('T27', 's'),
            ('U27', 't'),

            ('W27', 'd'),
            ('X27', 'o'),
            ('Y27', 'b'),
            ('Z27', 'u'),
            ('AA27', 't'),
            ('AB27', 's'),
            ('AC27', 'u'),

            ('AE27', 's'),
            ('AF27', 'h'),
            ('AG27', 'o'),
            ('AH27', 'g'),
            ('AI27', 'i'),
        ]

        for (cell_address, char) in input_data:
            cell = ws[cell_address]
            cell.value = char
            cell.font = self._TITLE_FONT

        cell = ws['L28']
        cell.value = 'Original 3x4 board was invented by Madoka Kitao. The original pieces were designed by Maiko Fujita.'
        cell.font = self._SMALL_TITLE_FONT
        cell.alignment = self._left_top_alignment
